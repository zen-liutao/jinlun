# -*- coding: UTF-8 -*-

import sys
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment


# import chardet
from datetime import datetime, timedelta


# 默认在表格c3位置 格式:2019/03/01 ~ 03/31	( 常州市金轮塑业 )
def extract_month(text):
    """
    抓取考勤的公司名称与开始日期
    :param text: 格式:2019/03/01 ~ 03/31	( 常州市金轮塑业 )
    :return: 开始的日期与公司名称
    """
    date = text.split()[0].strip()
    date = datetime.strptime(date, "%Y/%m/%d")
    # company_name = text.split()[3].strip()
    return date


def str2time(str_time):
    year, month, day = "2019", "01", "01"
    time = year + "-" + month + '-' + day + " " + str_time
    return datetime.strptime(time, '%Y-%m-%d %H:%M')
# todo 迟到早退的标准是什么，跟顾光宇商量下,扣钱的算法
# 暂时设置为距离上下班5分钟之内
# 短时间内多次重复刷卡
# 加班情况
# 迟到一个月给五次五分钟以内


def is_late(records_list):
    if not records_list:
        return False
    elif len(records_list) == 0:
        return False
    elif len(records_list) == 1:
        return False
    else:
        year, month, day = "2019", "01", "01"
        begin_time = records_list[0]
        end_time = records_list[-1]
        begin_time = str2time(begin_time)
        end_time = str2time(end_time)
        punch_in = str2time("7:30")
        punch_out = str2time("4:30")
        if punch_in < begin_time <= str2time("7:35") \
                or punch_out < end_time <= str2time("4:35"):
            return True
        return False


def get_attendance_time(lst, can_exempt=False):
    '''
    >>> get_attendance_time(['07:30','16:30'])
    1.0
    >>> get_attendance_time(['07:29','16:33'])
    1.0
    >>> get_attendance_time(['07:28','16:50'])
    1.0
    >>> get_attendance_time(['07:28','16:29'])
    0.9375
    >>> get_attendance_time(['07:44','17:25'])
    1.0
    >>> get_attendance_time(['07:24','19:10'])
    1.3125
    >>> get_attendance_time(['07:29','11:59',"16:31"])
    1.0
    '''

    year, month, day = "2019", "01", "01"

    def next_half_hour(time):
        if time.minute == 0 or time.minute == 30:
            return time
        time = time + timedelta(minutes=30)
        if time.minute < 30:
            time = time.replace(minute=0)
        else:
            time = time.replace(minute=30)
        return time

    def prev_half_hour(time):
        if time.minute == 0 or time.minute == 30:
            return time
        if time.minute > 30:
            time = time.replace(minute=30)
        else:
            time = time.replace(minute=0)
        return time

    def compute_work_hours(begin_time, end_time):
        begin_time = str2time(begin_time)
        end_time = str2time(end_time)
        begin_time = next_half_hour(begin_time)
        end_time = prev_half_hour(end_time)
        work_hours = (end_time - begin_time).seconds / 3600
        return work_hours

    if not lst:
        return 0
    elif can_exempt and is_late(lst):
        return 8
    else:
        if len(lst) == 0:
            return 0
        elif len(lst) == 1:
            return '异常'
        elif len(lst) == 2:
            begin_time, end_time = lst
            work_hours_morning = compute_work_hours(begin_time, '11:00')
            work_hours_afternoon = compute_work_hours('12:00', end_time)
            return work_hours_morning + work_hours_afternoon
        elif len(lst) == 3:
            begin_time1, begin_time2, end_time2 = lst
            work_hours_morning = compute_work_hours(begin_time1, '11:00')
            work_hours_afternoon = compute_work_hours(begin_time2, end_time2)
            return work_hours_morning + work_hours_afternoon
        elif len(lst) == 4:
            begin_time1, end_time1, begin_time2, end_time2 = lst
            work_hours_morning = compute_work_hours(begin_time1, end_time1)
            work_hours_afternoon = compute_work_hours(begin_time2, end_time2)
            return work_hours_morning + work_hours_afternoon
        else:
            return '异常'


def compute(sheet):
    # 获取表格的最大行与最大列
    max_row = sheet.max_row
    max_column = sheet.max_column
    # # 抽取开始日期
    attendance_date = extract_month(sheet['C3'].value)
    month = str(attendance_date.month)
    # begin_date = extract_month(attendance_date)[0]
    # # 抽取员工名字
    # employee_names = []
    # for i in range(5, max_row + 1, 2):
    #     employee_names.append(sheet.cell(i,11).value)

    d = {}
    # 抽取日期与工作时间
    for i in range(5, max_row + 1, 2):
          # 员工名字
        employee_name = sheet.cell(i, 11).value
        d[employee_name] = {'late_times': 0}
        for j in range(1, max_column+1):

            # 日期
            date = str(sheet.cell(4, j).value)
            # 工时
            work_hours = sheet.cell(i+1, j).value
            d[employee_name][month+"-"+date] = {}

            d[employee_name][month+"-"+date]['raw_work_hours'] = work_hours

            if work_hours:
                work_hours_list = work_hours.split()
                if is_late(work_hours_list):
                    d[employee_name]["late_times"] = d[employee_name]["late_times"] + 1
                if d[employee_name]["late_times"] > 5:
                    attendance_time = get_attendance_time(work_hours_list)
                else:
                    attendance_time = get_attendance_time(
                        work_hours_list, True)
                d[employee_name][month+"-"+date]['work_hours'] = attendance_time

    return d


# 把数据写入到新的表格中去
def writeSheet(data, wb):
    sheet = wb.worksheets[2]
    ws2 = wb.create_sheet("工时计算表", 0)   
    attendance_date = extract_month(sheet['C3'].value)
    ws2['a1'] = '常州市金轮塑业有限公司 考勤表{}年{}月'.format(str(attendance_date.year), str(attendance_date.month))
    ws2['a2'] = '日期/姓名'
    ws2['a2'].alignment = Alignment(horizontal="center", vertical="center")

    # 输入日期
    days = [day.split("-")[1]
            for day in data['顾伟文'].keys() if day != 'late_times']

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3+len(days)*2)
    ws2['a1'].alignment = Alignment(horizontal="center", vertical="center")
    i = 0
    for day in days:
        ws2.cell(row=2, column=2+2*i, value=day)
        ws2.merge_cells(start_row=2, start_column=2+2*i, end_row=2,end_column=3+2*i)
        # ws.merge_cells(start_row=2, start_column=2 + i, end_row=2, end_column=4 + i)
        i += 1
    ws2.cell(row=2, column=2 + len(days)*2, value="合计")
    ws2.cell(row=2, column=3 + len(days)*2, value="迟到次数")

    names = data.keys()
    i = 0
    days = [day for day in data['顾伟文'].keys() if day != 'late_times']
    for name in names:
        ws2.cell(row=3+i, column=1, value=name)
        i += 1
        employee_attendance = data[name]
        total_work_hours = 0
        j = 1
        for day in days:
            work_hours = employee_attendance[day].get("work_hours", 0)
            raw_work_hours = employee_attendance[day].get("raw_work_hours", 0)
            if type(work_hours) is float or type(work_hours) is int:
                work_hours = work_hours / 8
                total_work_hours += work_hours
            else:
                ws2.cell(row=2+i, column=2+j).fill = PatternFill(fgColor="FF0000", fill_type = "solid")
            ws2.cell(row=2+i, column=1+j, value=raw_work_hours)
            ws2.cell(row=2+i, column=1+j).alignment = Alignment(wrapText=True)
            ws2.cell(row=2+i, column=2+j, value=work_hours)
            j += 2
        ws2.cell(row=2+i, column=3 + len(days)*2, value=employee_attendance['late_times'])
        ws2.cell(row=2+i, column=2 + len(days)*2, value=total_work_hours)


def main():
    # print('参数个数为：', len(sys.argv))
    # print('参数列表:', str(sys.argv))
    # print('脚本名为: ', sys.argv[0])
    # for i in range(1, len(sys.argv)):
    #     print('参数 %s 为: %s' %(i, sys.argv[i]))
    # 从命令行获取excel文件路径
    file_path = sys.argv[1]
    # wb = load_workbook("05汇总表.xlsx")
    sheet = wb.worksheets[2]
    d = compute(sheet)
    writeSheet(d, wb)
    wb.save("结果：xxx.xlsx")


if __name__ == "__main__":
    main()
