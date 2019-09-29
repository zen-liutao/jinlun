# -*- coding: UTF-8 -*-

import sys
from openpyxl import Workbook
import datetime

def main():
    print('参数个数为：', len(sys.argv))
    print('参数列表:', str(sys.argv))
    print('脚本名为: ', sys.argv[0])
    for i in range(1, len(sys.argv)):
        print('参数 %s 为: %s' %(i, sys.argv[i]))
    # 从命令行获取excel文件路径
    file_path = sys.argv[1]
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 42
    ws.append([1, 2, 3])
    ws['A2'] = datetime.datetime.now()

    wb.save('sample.txt')

if __name__ == "__main__":
    main()