from openpyxl import Workbook
import datetime
   
wb = Workbook()
ws = wb.active
ws['A1'] = 42
ws.append([1, 2, 3])
ws['A2'] = datetime.datetime.now()

wb.save('sample.xlsx')